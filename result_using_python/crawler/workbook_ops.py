from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .config import FINAL_TYPES
from .models import CandidateResult, RowTask
from .text_utils import extract_brand


SCRIPT_DIR = Path(__file__).resolve().parent.parent


def resolve_user_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return (SCRIPT_DIR / path).resolve()


def get_sheet_and_headers(workbook, sheet_name: str | None):
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = {}
    for cell in worksheet[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return worksheet, headers


def ensure_output_columns(
    worksheet, type_col: int, headers: dict[str, int]
) -> tuple[int, int]:
    evidence_col = headers.get("Evidence URL")
    url_prove_col = headers.get("URL Prove")

    if evidence_col and url_prove_col:
        return evidence_col, url_prove_col

    insert_at = type_col + 1
    if not evidence_col and not url_prove_col:
        worksheet.insert_cols(insert_at, amount=2)
        worksheet.cell(row=1, column=insert_at, value="Evidence URL")
        worksheet.cell(row=1, column=insert_at + 1, value="URL Prove")
        return insert_at, insert_at + 1

    if not evidence_col:
        worksheet.insert_cols(insert_at)
        worksheet.cell(row=1, column=insert_at, value="Evidence URL")
        if url_prove_col and url_prove_col >= insert_at:
            url_prove_col += 1
        return insert_at, url_prove_col

    insert_at = evidence_col + 1
    worksheet.insert_cols(insert_at)
    worksheet.cell(row=1, column=insert_at, value="URL Prove")
    return evidence_col, insert_at


_HYBRID_CAR_INFO_KEYWORDS = frozenset([
    "hybrid", "phev", "mhev", "plug-in", "plug in", "recharge",
    "t8", "e-tron", "ioniq", "prius", "c-hr", "niro",
])


def _car_info_has_hybrid_hint(car_info: str) -> bool:
    lowered = car_info.lower()
    return any(kw in lowered for kw in _HYBRID_CAR_INFO_KEYWORDS)


def should_process_type(
    normalized_type: str,
    current_type,
    current_url,
    current_url_prove,
    car_info: str = "",
    *,
    overwrite: bool,
    allow_unknown: bool,
    fix_hybrid: bool = False,
) -> bool:
    # --fix-hybrid: reprocess electric/petrol rows that have no hybrid hint in car_info
    if fix_hybrid and normalized_type == "electric/petrol":
        if car_info and not _car_info_has_hybrid_hint(car_info):
            return True
        return False
    if normalized_type in FINAL_TYPES:
        return False
    if normalized_type == "unknown":
        return allow_unknown
    if normalized_type not in {"", "unknown"}:
        return False
    if not overwrite and current_type and current_url and current_url_prove is not None:
        return False
    return True


def build_row_task(row_number: int, car_info: str) -> RowTask:
    normalized_car_info = str(car_info).strip()
    return RowTask(
        row_number=row_number,
        car_info=normalized_car_info,
        brand=extract_brand(normalized_car_info),
    )


def iter_target_rows(
    worksheet,
    car_info_col: int,
    type_col: int,
    evidence_col: int,
    url_prove_col: int,
    limit: int,
    overwrite: bool,
    allow_unknown: bool,
    fix_hybrid: bool = False,
) -> Iterable[RowTask]:
    yielded = 0
    for row in range(2, worksheet.max_row + 1):
        car_info = worksheet.cell(row=row, column=car_info_col).value
        current_type = worksheet.cell(row=row, column=type_col).value
        current_url = worksheet.cell(row=row, column=evidence_col).value
        current_url_prove = worksheet.cell(row=row, column=url_prove_col).value
        normalized_type = str(current_type).strip().lower() if current_type else ""

        if not car_info:
            continue
        if not should_process_type(
            normalized_type,
            current_type,
            current_url,
            current_url_prove,
            car_info=str(car_info),
            overwrite=overwrite,
            allow_unknown=allow_unknown,
            fix_hybrid=fix_hybrid,
        ):
            continue

        yield build_row_task(row, car_info)
        yielded += 1
        if yielded >= limit:
            return


def load_workbook_context(args):
    input_path = resolve_user_path(args.input)
    output_path = resolve_user_path(args.output)

    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    workbook = load_workbook(input_path)
    worksheet, headers = get_sheet_and_headers(workbook, args.sheet)
    if "Car Info" not in headers or "Type" not in headers:
        raise SystemExit("The worksheet must contain 'Car Info' and 'Type' headers.")

    car_info_col = headers["Car Info"]
    type_col = headers["Type"]
    evidence_col, url_prove_col = ensure_output_columns(worksheet, type_col, headers)
    return workbook, worksheet, output_path, car_info_col, type_col, evidence_col, url_prove_col


def write_result_row(
    worksheet,
    output_path: Path,
    workbook,
    type_col: int,
    evidence_col: int,
    url_prove_col: int,
    task: RowTask,
    result: CandidateResult,
) -> None:
    worksheet.cell(row=task.row_number, column=type_col, value=result.fuel_type)
    worksheet.cell(row=task.row_number, column=evidence_col, value=result.url)
    worksheet.cell(row=task.row_number, column=url_prove_col, value=result.url_prove)
    workbook.save(output_path)
