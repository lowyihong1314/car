from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from .config import FINAL_TYPES
from .models import CandidateResult, RowTask
from .text_utils import extract_brand


SCRIPT_DIR = Path(__file__).resolve().parent.parent


def resolve_user_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return (SCRIPT_DIR / path).resolve()


def connect_database(db_path: str | Path) -> sqlite3.Connection:
    resolved_path = db_path if isinstance(db_path, Path) else resolve_user_path(str(db_path))
    connection = sqlite3.connect(resolved_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute("PRAGMA synchronous = NORMAL")
    return connection


def _connect(db_path: Path) -> sqlite3.Connection:
    return connect_database(db_path)


def _ensure_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS vehicles (
            row_number INTEGER PRIMARY KEY,
            car_info TEXT NOT NULL,
            brand TEXT NOT NULL,
            type TEXT,
            evidence_url TEXT,
            url_prove INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_vehicles_type ON vehicles(type);
        CREATE INDEX IF NOT EXISTS idx_vehicles_brand ON vehicles(brand);

        CREATE TABLE IF NOT EXISTS import_metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )
    connection.commit()


def _normalize_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_bool(value) -> int | None:
    if value is None or value == "":
        return None
    return int(bool(value))


def import_xlsx_into_db(db_path: Path, xlsx_path: Path) -> int:
    if not xlsx_path.exists():
        raise SystemExit(f"Bootstrap workbook not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    sheet_name = worksheet.title
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(value).strip(): index for index, value in enumerate(header_row) if value}
    required_headers = {"Car Info", "Type", "Evidence URL", "URL Prove"}
    missing_headers = sorted(required_headers - headers.keys())
    if missing_headers:
        raise SystemExit(
            f"Bootstrap workbook must contain headers {sorted(required_headers)}; missing {missing_headers}"
        )

    rows_to_insert: list[tuple[int, str, str, str | None, str | None, int | None]] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        car_info = _normalize_text(values[headers["Car Info"]])
        if not car_info:
            continue
        fuel_type = _normalize_text(values[headers["Type"]])
        evidence_url = _normalize_text(values[headers["Evidence URL"]])
        url_prove = _normalize_bool(values[headers["URL Prove"]])
        rows_to_insert.append(
            (
                row_number,
                car_info,
                extract_brand(car_info),
                fuel_type,
                evidence_url,
                url_prove,
            )
        )

    workbook.close()

    connection = _connect(db_path)
    try:
        connection.executescript(
            """
            DROP TABLE IF EXISTS vehicles;
            DROP TABLE IF EXISTS import_metadata;
            """
        )
        _ensure_schema(connection)
        connection.executemany(
            """
            INSERT INTO vehicles (
                row_number,
                car_info,
                brand,
                type,
                evidence_url,
                url_prove
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            rows_to_insert,
        )
        connection.executemany(
            "INSERT OR REPLACE INTO import_metadata(key, value) VALUES (?, ?)",
            [
                ("source_xlsx", str(xlsx_path)),
                ("sheet_name", sheet_name),
                ("row_count", str(len(rows_to_insert))),
            ],
        )
        connection.commit()
    finally:
        connection.close()

    return len(rows_to_insert)


_HYBRID_CAR_INFO_KEYWORDS = frozenset([
    "hybrid", "phev", "mhev", "plug-in", "plug in", "recharge",
    "t8", "e-tron", "ioniq", "prius", "c-hr", "niro",
])


def _car_info_has_hybrid_hint(car_info: str) -> bool:
    lowered = car_info.lower()
    return any(keyword in lowered for keyword in _HYBRID_CAR_INFO_KEYWORDS)


def should_process_type(
    normalized_type: str,
    current_type,
    current_url,
    current_url_prove,
    car_info: str = "",
    *,
    overwrite: bool,
    allow_unknown: bool,
    fix_hybrid: bool = False,
) -> bool:
    if fix_hybrid and normalized_type == "electric/petrol":
        if car_info and not _car_info_has_hybrid_hint(car_info):
            return True
        return False
    if normalized_type in FINAL_TYPES:
        return False
    if normalized_type == "unknown":
        return allow_unknown
    if normalized_type not in {"", "unknown"}:
        return False
    if not overwrite and current_type and current_url and current_url_prove is not None:
        return False
    return True


def load_database_context(args) -> tuple[sqlite3.Connection, Path, int | None, Path]:
    return load_database_context_values(
        db=args.db,
        bootstrap_xlsx=args.bootstrap_xlsx,
        reimport_db=args.reimport_db,
    )


def load_database_context_values(
    *,
    db: str,
    bootstrap_xlsx: str,
    reimport_db: bool,
) -> tuple[sqlite3.Connection, Path, int | None, Path]:
    db_path = resolve_user_path(db)
    bootstrap_xlsx_path = resolve_user_path(bootstrap_xlsx)

    imported_rows: int | None = None
    if reimport_db or not db_path.exists():
        db_path.parent.mkdir(parents=True, exist_ok=True)
        imported_rows = import_xlsx_into_db(db_path, bootstrap_xlsx_path)

    connection = _connect(db_path)
    _ensure_schema(connection)
    return connection, db_path, imported_rows, bootstrap_xlsx_path


def iter_target_rows(
    connection: sqlite3.Connection,
    limit: int,
    overwrite: bool,
    allow_unknown: bool,
    fix_hybrid: bool = False,
    start_row: int = 1,
):
    yielded = 0
    safe_start_row = max(1, int(start_row))
    cursor = connection.execute(
        """
        SELECT row_number, car_info, brand, type, evidence_url, url_prove
        FROM vehicles
        WHERE row_number >= ?
        ORDER BY row_number
        """,
        (safe_start_row,),
    )

    for row in cursor:
        car_info = str(row["car_info"]).strip() if row["car_info"] else ""
        if not car_info:
            continue

        current_type = row["type"]
        current_url = row["evidence_url"]
        current_url_prove = None if row["url_prove"] is None else bool(row["url_prove"])
        normalized_type = str(current_type).strip().lower() if current_type else ""

        if not should_process_type(
            normalized_type,
            current_type,
            current_url,
            current_url_prove,
            car_info=car_info,
            overwrite=overwrite,
            allow_unknown=allow_unknown,
            fix_hybrid=fix_hybrid,
        ):
            continue

        yield RowTask(
            row_number=int(row["row_number"]),
            car_info=car_info,
            brand=str(row["brand"]).strip() if row["brand"] else extract_brand(car_info),
        )
        yielded += 1
        if yielded >= limit:
            return


def write_result_row(
    connection: sqlite3.Connection,
    task: RowTask,
    result: CandidateResult,
) -> None:
    connection.execute(
        """
        UPDATE vehicles
        SET type = ?, evidence_url = ?, url_prove = ?, updated_at = CURRENT_TIMESTAMP
        WHERE row_number = ?
        """,
        (
            result.fuel_type,
            result.url or None,
            int(bool(result.url_prove)),
            task.row_number,
        ),
    )
    connection.commit()


def fetch_vehicle_rows(
    connection: sqlite3.Connection,
    *,
    search: str = "",
    fuel_type: str = "",
    limit: int = 100,
    offset: int = 0,
) -> tuple[list[dict], int]:
    normalized_search = search.strip().lower()
    normalized_type = fuel_type.strip().lower()
    safe_limit = max(1, min(int(limit), 30))
    safe_offset = max(0, int(offset))

    where_clauses: list[str] = []
    params: list[object] = []

    if normalized_search:
        where_clauses.append(
            """
            (
                CAST(row_number AS TEXT) LIKE ?
                OR lower(car_info) LIKE ?
                OR lower(brand) LIKE ?
                OR lower(COALESCE(type, '')) LIKE ?
                OR lower(COALESCE(evidence_url, '')) LIKE ?
            )
            """
        )
        like_pattern = f"%{normalized_search}%"
        params.extend([like_pattern] * 5)

    if normalized_type:
        where_clauses.append("lower(COALESCE(type, '')) = ?")
        params.append(normalized_type)

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    total = connection.execute(
        f"SELECT COUNT(*) FROM vehicles {where_sql}",
        params,
    ).fetchone()[0]

    rows = connection.execute(
        f"""
        SELECT
            row_number,
            car_info,
            brand,
            type,
            evidence_url,
            url_prove,
            updated_at
        FROM vehicles
        {where_sql}
        ORDER BY row_number
        LIMIT ? OFFSET ?
        """,
        [*params, safe_limit, safe_offset],
    ).fetchall()

    payload_rows = [
        {
            "row_number": int(row["row_number"]),
            "car_info": row["car_info"],
            "brand": row["brand"],
            "type": row["type"] or "",
            "evidence_url": row["evidence_url"] or "",
            "url_prove": None if row["url_prove"] is None else bool(row["url_prove"]),
            "updated_at": row["updated_at"],
        }
        for row in rows
    ]
    return payload_rows, int(total)


def fetch_dashboard_summary(connection: sqlite3.Connection) -> dict[str, int]:
    row = connection.execute(
        """
        SELECT
            COUNT(*) AS total_rows,
            SUM(CASE WHEN type = 'diesel' THEN 1 ELSE 0 END) AS diesel_count,
            SUM(CASE WHEN type = 'petrol' THEN 1 ELSE 0 END) AS petrol_count,
            SUM(CASE WHEN type = 'electric' THEN 1 ELSE 0 END) AS electric_count,
            SUM(CASE WHEN type = 'electric/petrol' THEN 1 ELSE 0 END) AS hybrid_count,
            SUM(CASE WHEN type = 'unknown' THEN 1 ELSE 0 END) AS unknown_count,
            SUM(CASE WHEN type IS NULL OR trim(type) = '' THEN 1 ELSE 0 END) AS empty_count,
            SUM(CASE WHEN url_prove = 1 THEN 1 ELSE 0 END) AS proven_count
        FROM vehicles
        """
    ).fetchone()
    return {key: int(row[key] or 0) for key in row.keys()}


def fetch_brand_type_breakdown(connection: sqlite3.Connection) -> list[dict]:
    rows = connection.execute(
        """
        SELECT
            brand,
            COUNT(*) AS total_rows,
            SUM(CASE WHEN type = 'diesel' THEN 1 ELSE 0 END) AS diesel_count,
            SUM(CASE WHEN type = 'petrol' THEN 1 ELSE 0 END) AS petrol_count,
            SUM(CASE WHEN type = 'electric' THEN 1 ELSE 0 END) AS electric_count,
            SUM(CASE WHEN type = 'electric/petrol' THEN 1 ELSE 0 END) AS hybrid_count,
            SUM(CASE WHEN type = 'unknown' THEN 1 ELSE 0 END) AS unknown_count,
            SUM(CASE WHEN type IS NULL OR trim(type) = '' THEN 1 ELSE 0 END) AS empty_count
        FROM vehicles
        WHERE trim(COALESCE(brand, '')) <> ''
        GROUP BY brand
        ORDER BY total_rows DESC, brand ASC
        """
    ).fetchall()
    return [
        {
            "brand": row["brand"],
            "total_rows": int(row["total_rows"] or 0),
            "diesel_count": int(row["diesel_count"] or 0),
            "petrol_count": int(row["petrol_count"] or 0),
            "electric_count": int(row["electric_count"] or 0),
            "hybrid_count": int(row["hybrid_count"] or 0),
            "unknown_count": int(row["unknown_count"] or 0),
            "empty_count": int(row["empty_count"] or 0),
        }
        for row in rows
    ]


def fetch_recent_updates(
    connection: sqlite3.Connection,
    *,
    limit: int = 20,
) -> list[dict]:
    safe_limit = max(1, min(int(limit), 100))
    rows = connection.execute(
        """
        SELECT row_number, car_info, type, evidence_url, url_prove, updated_at
        FROM vehicles
        ORDER BY datetime(updated_at) DESC, row_number DESC
        LIMIT ?
        """,
        (safe_limit,),
    ).fetchall()
    return [
        {
            "row_number": int(row["row_number"]),
            "car_info": row["car_info"],
            "type": row["type"] or "",
            "evidence_url": row["evidence_url"] or "",
            "url_prove": None if row["url_prove"] is None else bool(row["url_prove"]),
            "updated_at": row["updated_at"],
        }
        for row in rows
    ]
