#!/usr/bin/env python3
"""Fill fuel types first, then evidence URLs, with per-row autosave.

This script reads the first worksheet in an Excel file, looks for the
`Car Info` and `Type` columns, and writes `Evidence URL` plus `URL Prove`
next to `Type`. It processes at most 10 rows per run by default.

The OpenAI Responses API is used in two stages:
1. classify fuel type in brand-based batches
2. search for a supporting URL per row

The workbook is saved after each stage for each row.
"""

from __future__ import annotations

import argparse
import json
import os
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


API_URL = "https://api.openai.com/v1/responses"
DEFAULT_MODEL = "gpt-5"
DEFAULT_BATCH_SIZE = 10
DEFAULT_BRAND_BATCHES = 10
OUTPUT_FILENAME = "carlist_type_mapping_ai.xlsx"
ALLOWED_TYPES = ["diesel", "petrol", "electric", "electric/petrol", "unknown"]
FINAL_TYPES = {"diesel", "petrol", "electric", "electric/petrol"}
SCRIPT_DIR = Path(__file__).resolve().parent
ENV_FILE = SCRIPT_DIR / ".env"


@dataclass
class RowTask:
    row_number: int
    car_info: str
    brand: str


@dataclass
class BrandBatch:
    brand: str
    rows: list[RowTask]


def load_dotenv_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()

        if not key:
            continue
        if value.startswith(("\"", "'")) and value.endswith(("\"", "'")):
            value = value[1:-1]

        os.environ.setdefault(key, value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fill the Type, Evidence URL, and URL Prove columns in "
            "'Carlist Type Mapping.xlsx' using the OpenAI API."
        )
    )
    parser.add_argument(
        "--input",
        default="carlist_type_mapping_ai.xlsx",
        help="Path to the source workbook.",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT_FILENAME,
        help="Path to the output workbook.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="Maximum models per brand batch.",
    )
    parser.add_argument(
        "--brand-batches",
        type=int,
        default=DEFAULT_BRAND_BATCHES,
        help="Maximum number of brand batches to process in one run.",
    )
    parser.add_argument(
        "--model",
        default=os.getenv("OPENAI_MODEL", DEFAULT_MODEL),
        help="OpenAI model name. Defaults to OPENAI_MODEL or gpt-5.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Re-process rows even if Type, Evidence URL, and URL Prove are already filled.",
    )
    parser.add_argument(
        "--allow-unknown",
        action="store_true",
        help="Also process rows where Type is 'unknown'. By default only empty Type rows are processed.",
    )
    return parser.parse_args()


def require_api_key() -> str:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise SystemExit("OPENAI_API_KEY is not set.")
    return api_key


def resolve_user_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return (SCRIPT_DIR / path).resolve()


def extract_brand(car_info: str) -> str:
    return car_info.split(" ", 1)[0].strip()


def build_type_batch_prompt(brand: str, rows: list[RowTask]) -> str:
    models_text = "\n".join(
        f"- row_number: {row.row_number}, car_info: {row.car_info}" for row in rows
    )
    return f"""
You are classifying vehicles from the same brand into fuel type buckets.

Return JSON only, matching the provided schema.

Brand:
{brand}

Rows:
{models_text}

Rules:
- Return one result for every input row_number
- Choose exactly one `type` per row from: diesel, petrol, electric, electric/petrol, unknown
- `electric/petrol` is for hybrids only
- `reason` must be brief, 1 sentence max for each row
- Use your general vehicle knowledge and the model name only
- Do not search the web in this step
- If you are not reasonably confident, return `unknown`
""".strip()


def build_evidence_prompt(car_info: str, fuel_type: str) -> str:
    return f"""
Find one public web page that supports this fuel type classification.

Return JSON only, matching the provided schema.

Vehicle name:
{car_info}

Chosen type:
{fuel_type}

Rules:
- Search the web for supporting evidence
- Return a direct page URL if you find support
- Set `url_prove` to true only if the page clearly supports the chosen type
- Set `url_prove` to false if the page is ambiguous, unrelated, or contradicts the chosen type
- Set `url_prove` to false if you cannot find a reliable supporting URL
- `reason` must be brief, 1 sentence max
""".strip()


def call_openai_json(
    api_key: str,
    model: str,
    prompt: str,
    *,
    schema_name: str,
    schema: dict,
    use_web_search: bool,
) -> dict:
    payload = {
        "model": model,
        "reasoning": {"effort": "low"},
        "text": {
            "format": {
                "type": "json_schema",
                "name": schema_name,
                "strict": True,
                "schema": schema,
            }
        },
        "input": prompt,
    }
    if use_web_search:
        payload["tools"] = [{"type": "web_search"}]
        payload["tool_choice"] = "required"
        payload["include"] = ["web_search_call.action.sources"]

    request = urllib.request.Request(
        API_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"OpenAI API error {exc.code}: {body}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"OpenAI API request failed: {exc}") from exc


def extract_response_json(response: dict) -> dict:
    for item in response.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") != "output_text":
                continue
            text = content.get("text", "").strip()
            if not text:
                continue
            return json.loads(text)
    raise RuntimeError("No structured JSON content found in OpenAI response.")


def extract_source_urls(response: dict) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()

    def maybe_add(url: str | None) -> None:
        if not url:
            return
        normalized = url.strip()
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        urls.append(normalized)

    for item in response.get("output", []):
        if item.get("type") != "web_search_call":
            continue
        action = item.get("action", {})
        for source in action.get("sources", []):
            maybe_add(source.get("url"))

    return urls


def extract_result(response: dict) -> dict:
    result = extract_response_json(response)
    source_urls = extract_source_urls(response)
    candidate_url = str(result.get("evidence_url", "")).strip()
    if source_urls and candidate_url not in source_urls:
        result["evidence_url"] = source_urls[0]
    elif not candidate_url and source_urls:
        result["evidence_url"] = source_urls[0]
    return result


def classify_type_batch(
    api_key: str, model: str, brand: str, rows: list[RowTask]
) -> list[dict]:
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "row_number": {"type": "integer"},
                        "type": {"type": "string", "enum": ALLOWED_TYPES},
                        "reason": {"type": "string"},
                    },
                    "required": ["row_number", "type", "reason"],
                },
            }
        },
        "required": ["results"],
    }
    response = call_openai_json(
        api_key,
        model,
        build_type_batch_prompt(brand, rows),
        schema_name="fuel_type_batch",
        schema=schema,
        use_web_search=False,
    )
    result = extract_response_json(response)
    return result["results"]


def find_evidence_url(api_key: str, model: str, car_info: str, fuel_type: str) -> dict:
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "evidence_url": {"type": "string"},
            "url_prove": {"type": "boolean"},
            "reason": {"type": "string"},
        },
        "required": ["evidence_url", "url_prove", "reason"],
    }
    response = call_openai_json(
        api_key,
        model,
        build_evidence_prompt(car_info, fuel_type),
        schema_name="fuel_type_evidence",
        schema=schema,
        use_web_search=True,
    )
    result = extract_result(response)
    evidence_url = str(result.get("evidence_url", "")).strip()
    if not evidence_url:
        result["url_prove"] = False
    elif not bool(result.get("url_prove", False)):
        result["evidence_url"] = evidence_url
    return result


def get_sheet_and_headers(workbook, sheet_name: str | None):
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = {}
    for cell in worksheet[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return worksheet, headers


def ensure_output_columns(
    worksheet, type_col: int, headers: dict[str, int]
) -> tuple[int, int]:
    evidence_col = headers.get("Evidence URL")
    url_prove_col = headers.get("URL Prove")

    if evidence_col and url_prove_col:
        return evidence_col, url_prove_col

    insert_at = type_col + 1
    if not evidence_col and not url_prove_col:
        worksheet.insert_cols(insert_at, amount=2)
        worksheet.cell(row=1, column=insert_at, value="Evidence URL")
        worksheet.cell(row=1, column=insert_at + 1, value="URL Prove")
        return insert_at, insert_at + 1

    if not evidence_col:
        worksheet.insert_cols(insert_at)
        worksheet.cell(row=1, column=insert_at, value="Evidence URL")
        if url_prove_col and url_prove_col >= insert_at:
            url_prove_col += 1
        return insert_at, url_prove_col

    insert_at = evidence_col + 1
    worksheet.insert_cols(insert_at)
    worksheet.cell(row=1, column=insert_at, value="URL Prove")
    return evidence_col, insert_at


def iter_target_rows(
    worksheet,
    car_info_col: int,
    type_col: int,
    evidence_col: int,
    url_prove_col: int,
    limit: int,
    overwrite: bool,
    allow_unknown: bool,
) -> Iterable[RowTask]:
    yielded = 0
    for row in range(2, worksheet.max_row + 1):
        car_info = worksheet.cell(row=row, column=car_info_col).value
        current_type = worksheet.cell(row=row, column=type_col).value
        current_url = worksheet.cell(row=row, column=evidence_col).value
        current_url_prove = worksheet.cell(row=row, column=url_prove_col).value
        normalized_type = str(current_type).strip().lower() if current_type else ""

        if not car_info:
            continue
        if normalized_type in FINAL_TYPES:
            continue
        if normalized_type == "unknown" and not allow_unknown:
            continue
        if normalized_type not in {"", "unknown"}:
            continue
        if not overwrite and current_type and current_url and current_url_prove is not None:
            continue

        normalized_car_info = str(car_info).strip()
        yield RowTask(
            row_number=row,
            car_info=normalized_car_info,
            brand=extract_brand(normalized_car_info),
        )
        yielded += 1
        # limit is enforced later as batch size per brand


def group_into_brand_batches(
    tasks: Iterable[RowTask], batch_size: int, max_brand_batches: int
) -> list[BrandBatch]:
    grouped: dict[str, list[RowTask]] = {}
    order: list[str] = []

    for task in tasks:
        if task.brand not in grouped:
            grouped[task.brand] = []
            order.append(task.brand)
        grouped[task.brand].append(task)

    batches: list[BrandBatch] = []
    for brand in order:
        brand_rows = grouped[brand]
        for start in range(0, len(brand_rows), batch_size):
            chunk = brand_rows[start : start + batch_size]
            if not chunk:
                continue
            batches.append(BrandBatch(brand=brand, rows=chunk))
            if len(batches) >= max_brand_batches:
                return batches
    return batches


def main() -> int:
    load_dotenv_file(ENV_FILE)
    args = parse_args()
    api_key = require_api_key()

    input_path = resolve_user_path(args.input)
    output_path = resolve_user_path(args.output)

    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    workbook = load_workbook(input_path)
    worksheet, headers = get_sheet_and_headers(workbook, args.sheet)

    if "Car Info" not in headers or "Type" not in headers:
        raise SystemExit("The worksheet must contain 'Car Info' and 'Type' headers.")

    car_info_col = headers["Car Info"]
    type_col = headers["Type"]
    evidence_col, url_prove_col = ensure_output_columns(worksheet, type_col, headers)

    tasks = list(
        iter_target_rows(
            worksheet=worksheet,
            car_info_col=car_info_col,
            type_col=type_col,
            evidence_col=evidence_col,
            url_prove_col=url_prove_col,
            limit=args.limit,
            overwrite=args.overwrite,
            allow_unknown=args.allow_unknown,
        )
    )

    if not tasks:
        print("No eligible rows found.")
        workbook.save(output_path)
        print(f"Saved workbook without changes: {output_path}")
        return 0

    brand_batches = group_into_brand_batches(
        tasks,
        batch_size=max(args.limit, 1),
        max_brand_batches=max(args.brand_batches, 1),
    )

    if not brand_batches:
        print("No eligible brand batches found.")
        workbook.save(output_path)
        print(f"Saved workbook without changes: {output_path}")
        return 0

    total_rows = sum(len(batch.rows) for batch in brand_batches)
    print(
        f"Processing {len(brand_batches)} brand batch(es), "
        f"{total_rows} row(s), model {args.model}..."
    )
    for batch_index, batch in enumerate(brand_batches, start=1):
        print(
            f"Brand batch {batch_index}/{len(brand_batches)}: "
            f"{batch.brand} ({len(batch.rows)} row(s))"
        )
        batch_results = classify_type_batch(api_key, args.model, batch.brand, batch.rows)
        batch_result_map = {
            int(item["row_number"]): item for item in batch_results if "row_number" in item
        }

        for task in batch.rows:
            print(f"Row {task.row_number}: {task.car_info}")
            current_type = worksheet.cell(row=task.row_number, column=type_col).value
            current_url = worksheet.cell(row=task.row_number, column=evidence_col).value
            current_url_prove = worksheet.cell(row=task.row_number, column=url_prove_col).value

            fuel_type = str(current_type).strip() if current_type else ""
            if args.overwrite or not fuel_type:
                type_result = batch_result_map.get(task.row_number)
                if not type_result:
                    fuel_type = "unknown"
                else:
                    fuel_type = str(type_result["type"]).strip()
                worksheet.cell(row=task.row_number, column=type_col, value=fuel_type)
                workbook.save(output_path)
                print(f"  -> saved type={fuel_type}")
            else:
                print(f"  -> keep existing type={fuel_type}")

            if not fuel_type:
                fuel_type = "unknown"

            needs_evidence = args.overwrite or not current_url or current_url_prove is None
            if not needs_evidence:
                print(f"  -> keep existing url={current_url}, url_prove={current_url_prove}")
                continue

            evidence_result = find_evidence_url(api_key, args.model, task.car_info, fuel_type)
            evidence_url = str(evidence_result.get("evidence_url", "")).strip()
            url_prove = bool(evidence_result.get("url_prove", False))
            worksheet.cell(row=task.row_number, column=evidence_col, value=evidence_url)
            worksheet.cell(row=task.row_number, column=url_prove_col, value=url_prove)
            workbook.save(output_path)
            print(f"  -> saved url={evidence_url or '-'}, url_prove={url_prove}")

    workbook.save(output_path)
    print(f"Saved: {output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit(130)
